#!/usr/bin/env python3
"""申請書類フォームのExcel入力セルを検出する（ハマり④）。

入力欄は「空だとピンク」＝条件付き書式 containsBlanks(LEN(TRIM(cell))=0)。
そのCF範囲＝入力セル。テンプレでは空なので、CF範囲→結合アンカー→
ラベル(テンプレに値ありの左見出し)を除外して抽出する。

使い方: TEMPLATE を実テンプレのパスに書き換えて実行。
  未マッピング入力欄だけ見たい時は MAPPED_CODES / MAPPED_CELLS を渡して除外する。
"""
import json
import re
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter, coordinate_to_tuple

# TODO: 実テンプレのパスに変更
TEMPLATE = "template.xlsx"
# 企業が実際に記入する表示シートだけ対象にする（DATA/dekisugi用/1-23等の非表示は除外）
INPUT_SHEETS = ["はじめに", "1-4", "1-6", "1-6別紙", "居住費の詳細",
                "【介護分野】事業所概要1"]
# app34(法人マスタ=実在企業)に書くセルは仮データ禁止。マッピングに応じて調整。
APP34_CELLS = set()  # 例: {("はじめに","E14"), ("1-11-1","K32"), ...}


def label_left(ws, ref):
    r, c = coordinate_to_tuple(ref)
    for cc in range(c - 1, max(0, c - 8), -1):
        v = ws.cell(row=r, column=cc).value
        if v not in (None, "") and not str(v).startswith("="):
            return str(v).replace("\n", " ").strip()[:32]
    return "?"


def anchor_of(ws, ref):
    r, c = coordinate_to_tuple(ref)
    for m in ws.merged_cells.ranges:
        mc, mr, xc, xr = range_boundaries(str(m))
        if mr <= r <= xr and mc <= c <= xc:
            return f"{get_column_letter(mc)}{mr}"
    return ref


def detect(template=TEMPLATE, mapped=frozenset()):
    wb = openpyxl.load_workbook(template)
    out = {}
    for s in INPUT_SHEETS:
        if s not in wb.sheetnames:
            continue
        ws = wb[s]
        anchors = {}
        for cf_range in ws.conditional_formatting:
            for rule in ws.conditional_formatting[cf_range]:
                if rule.type != "containsBlanks":
                    continue
                for part in str(cf_range.sqref).split():
                    mc, mr, xc, xr = range_boundaries(part)
                    for r in range(mr, xr + 1):
                        for c in range(mc, xc + 1):
                            ref = f"{get_column_letter(c)}{r}"
                            a = anchor_of(ws, ref)
                            if ws[a].value not in (None, ""):
                                continue  # ラベルは入力欄でない
                            if (s, a) in APP34_CELLS or (s, a) in mapped:
                                continue
                            anchors[a] = label_left(ws, a)
        out[s] = anchors
    return out


if __name__ == "__main__":
    res = detect()
    for sheet, cells in res.items():
        print(f"\n### {sheet} （{len(cells)}欄） ###")
        for a, lab in cells.items():
            print(f"  {a}: {lab}")
    print("\n合計:", sum(len(v) for v in res.values()))
